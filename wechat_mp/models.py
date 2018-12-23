# -*- coding: utf-8 -*-
"""
@version: python3.6
@author: "Roger Lee"
@license: MIT Licence 
@contact: 704480843@qq.com
@file: models.py
@time: 2018/9/8 11:16
"""
import random
import time
import logging

import openpyxl
from tqdm import tqdm

from wechat_mp.utils import from_timestamp_to_datetime_string
from wechat_mp.exceptions import ArticlesNotObtainError

logger = logging.getLogger('wechat_mp')


class OfficalAccount:
    """
    微信公众号的对象

    """
    def __init__(self, raw_dict, client):
        self.client = client
        self.fakeid = raw_dict.get('fakeid')
        self.nickname = raw_dict.get('nickname')
        self.alias = raw_dict.get('alias')
        self.round_head_img = raw_dict.get('round_head_img')
        self.service_type = raw_dict.get('service_type')

    def articles(self, title_contain="", limit=0):
        """
        获取公众号的历史群发图文

        :param title_contain: 图文标题包含的字符串
        :type title_contain: str
        :param limit: 限制获取的图文数量
        :type limit: int
        :return: :class:`Article <wechat_mp.models.Article>` 对象列表
        """
        search_api = self.client.api_collections('search', 'article list').format(self.client.token, random.randint(200, 999))

        params = {
            'begin': 0,
            'count': 5,
            'query': title_contain,
            'fakeid': self.fakeid
        }

        response = self.client.session.get(search_api, params=params).json()
        article_list = []
        begin = 0

        total = response.get('app_msg_cnt')

        if limit == 0:
            logger.info("该公众号共有%s篇图文,已设置不限制获取数量", total)
            bar = tqdm(total=total)
            limit = total
        else:
            logger.info("该公众号共有%s篇图文,已设置限制获取前%s篇图文", total, limit)
            bar = tqdm(total=limit)

        bar.set_description("进度条")
        reach_limit = False

        while not reach_limit:

            page_result = self._search_article_pages(search_api, params)
            print(len(page_result))
            if not page_result:
                reach_limit = True

            for article in page_result:
                if len(article_list) >= limit or len(article_list) > total:
                    reach_limit = True
                else:
                    article_list.append(article)
                    bar.update(1)
            begin += 5
            params['begin'] = begin
            time.sleep(3)

        article_result = ArticleSearchResult([Article(article) for article in article_list], type=1)
        bar.close()
        return article_result

    def _search_article_pages(self,search_api, params):
        """
        根据页数不断地进行请求

        :param api: 请求的API
        :param params: 包含起始的参数
        :return: 文章字典列表
        """
        articles = []
        response = self.client.session.get(search_api, params=params).json()
        if response['base_resp']['ret'] == 0:
            app_msg_list = response.get('app_msg_list')
            articles += app_msg_list
        return articles


    def __str__(self):
        return f"<{self.__class__.__name__}: {self.nickname}>"

    def __repr__(self):
        return f"<{self.__class__.__name__}: {self.nickname}>"

class Article:
    """
    微信图文对象
    """
    def __init__(self,raw_dict):
        self.aid = raw_dict.get('aid')
        self.appmsgid = raw_dict.get('appmsgid')
        self.cover = raw_dict.get('cover')
        self.digest = raw_dict.get('digest')
        self.itemidx = raw_dict.get('itemidx')
        self.link = raw_dict.get('link')
        self.title = raw_dict.get('title')
        self._update_time = raw_dict.get('update_time')

    @property
    def update_time(self):
        """
        将时间戳转成日期字符串

        :return: 日期字符串
        """
        return from_timestamp_to_datetime_string(int(self._update_time))

    def __str__(self):
        return f"<{self.__class__.__name__}: {self.title}>"

    def __repr__(self):
        return f"<{self.__class__.__name__}: {self.title}>"

class ArticleWithContent:
    """
    通过关键词搜索出来的图文对象
    比通过账号搜索出来的图文对象
    具有更多的图文属性
    """
    def __init__(self, raw_dict):
        self.article_type = raw_dict.get("article_type")
        self.author = raw_dict.get("author")
        self.content = raw_dict.get("content")
        self.cover_url = raw_dict.get("cover_url")
        self.head_img_url = raw_dict.get("head_img_url")
        self.nickname = raw_dict.get("nickname")
        self.source_can_reward = raw_dict.get("source_can_reward")
        self.source_reprint_status = raw_dict.get("source_reprint_status")
        self.source_url = raw_dict.get("source_url")
        self.title = raw_dict.get("title")
        self.url = raw_dict.get("url")

    def __str__(self):
        return f"<{self.__class__.__name__}: {self.title}>"

    def __repr__(self):
        return f"<{self.__class__.__name__}: {self.title}>"

class ArticleSearchResult:

    def __init__(self, article_list, type):
        self.article_list = article_list
        self.type = type

    @property
    def total(self):
        return len(self.article_list)

    def save_articles_as_excel(self,filename, include_content=False):
        """
        将该公众号搜索出来的历史群发图文保存到Excel中

        :param filename: 文件名。默认添加了.xlsx后缀
        :return: None
        """

        if self.type == 0:
            if not self.article_list:
                raise ArticlesNotObtainError(f"结果为空")

            def write_header(sheet):
                if include_content:
                    sheet.cell(1, 1, value="公众号名称")
                    sheet.cell(1, 2, value="标题")
                    sheet.cell(1, 3, value="作者")
                    sheet.cell(1, 4, value="正文")
                    sheet.cell(1, 5, value="图文地址")
                    sheet.cell(1, 6, value="封面地址")
                    sheet.cell(1, 7, value="公众号头像地址")
                else:
                    sheet.cell(1, 1, value="公众号名称")
                    sheet.cell(1, 2, value="标题")
                    sheet.cell(1, 3, value="作者")
                    sheet.cell(1, 4, value="图文地址")
                    sheet.cell(1, 5, value="封面地址")
                    sheet.cell(1, 6, value="公众号头像地址")

                return sheet

            if 'xlsx' not in filename:
                filename += '.xlsx'
            wb = openpyxl.Workbook()
            wb.save(filename)
            wb = openpyxl.load_workbook(filename)
            sheet = wb[wb.sheetnames[0]]
            sheet.title = "图文列表"
            sheet = write_header(sheet)

            for index, article in enumerate(self.article_list,2):
                if include_content:
                    sheet.cell(index, 1, value=article.nickname)
                    sheet.cell(index, 2, value=article.title)
                    sheet.cell(index, 3, value=article.author)
                    sheet.cell(index, 4, value=article.content)
                    sheet.cell(index, 5, value=article.url)
                    sheet.cell(index, 6, value=article.cover_url)
                    sheet.cell(index, 7, value=article.head_img_url)
                else:
                    sheet.cell(index, 1, value=article.nickname)
                    sheet.cell(index, 2, value=article.title)
                    sheet.cell(index, 3, value=article.author)
                    sheet.cell(index, 4, value=article.url)
                    sheet.cell(index, 5, value=article.cover_url)
                    sheet.cell(index, 6, value=article.head_img_url)

            wb.save(filename)
        elif self.type == 1:
            def write_header(sheet):
                sheet.cell(1, 1, value="标题")
                sheet.cell(1, 2, value="摘要")
                sheet.cell(1, 3, value="链接")
                sheet.cell(1, 4, value="更新时间")
                sheet.cell(1, 5, value="aid")
                sheet.cell(1, 6, value="appmsgid")
                sheet.cell(1, 7, value="图文序号")
                return sheet

            if 'xlsx' not in filename:
                filename += '.xlsx'
            wb = openpyxl.Workbook()
            wb.save(filename)
            wb = openpyxl.load_workbook(filename)
            sheet = wb[wb.sheetnames[0]]
            sheet.title = "图文列表"
            sheet = write_header(sheet)

            for index, article in enumerate(self.article_list, 2):
                sheet.cell(index, 1, value=article.title)
                sheet.cell(index, 2, value=article.digest)
                sheet.cell(index, 3, value=article.link)
                sheet.cell(index, 4, value=article.update_time)
                sheet.cell(index, 5, value=article.aid)
                sheet.cell(index, 6, value=article.appmsgid)
                sheet.cell(index, 7, value=article.itemidx)

            wb.save(filename)

    def __getitem__(self, item):
        if isinstance(item, int):
            return self.article_list[item]


    def __str__(self):
        return f"<{self.__class__.__name__}: {len(self.article_list)}>"

    def __repr__(self):
        return f"<{self.__class__.__name__}: {len(self.article_list)}>"


